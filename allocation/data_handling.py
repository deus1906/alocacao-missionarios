"""Input parsing and validation for the allocation model."""

from __future__ import annotations

import io
import math
import re
import unicodedata
from typing import Any, Dict, Iterable, List, Tuple

from loguru import logger

from allocation.config import VALENCIAS_SHEETS, MISSIONARIOS_SHEETS


#=========================================================================
# 0. DATA LOADING
#=========================================================================

def _strip_accents(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _normalize_key(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = _strip_accents(text)
    text = text.lower().strip()
    return re.sub(r"\s+", "", text)


def _normalize_label(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return " ".join(text.split())


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    if isinstance(value, float):
        return math.isnan(value)
    return False


def _load_workbook(source: Any) -> Tuple[str, Any]:
    try:
        import pandas as pd

        if isinstance(source, (bytes, bytearray)):
            return "pandas", pd.ExcelFile(io.BytesIO(source))
        return "pandas", pd.ExcelFile(source)
    except Exception as pandas_exc:
        logger.debug("Pandas load failed: {}", pandas_exc)
        try:
            import openpyxl

            if isinstance(source, (bytes, bytearray)):
                return "openpyxl", openpyxl.load_workbook(
                    io.BytesIO(source), data_only=True
                )
            return "openpyxl", openpyxl.load_workbook(source, data_only=True)
        except Exception as openpyxl_exc:
            raise RuntimeError(
                "Failed to load Excel file. Install pandas or openpyxl."
            ) from openpyxl_exc


def _find_sheet(sheet_names: Iterable[str], candidates: Iterable[str]) -> str:
    normalized = {_normalize_key(name): name for name in sheet_names}
    for candidate in candidates:
        key = _normalize_key(candidate)
        if key in normalized:
            return normalized[key]
    if len(sheet_names) == 1:
        return sheet_names[0]
    available = ", ".join(sheet_names)
    raise ValueError(
        f"Could not find a matching sheet name. Available sheets: {available}"
    )



def _read_sheet_records(mode: str, book: Any, sheet_name: str) -> Tuple[List[Dict[str, Any]], List[Any]]:
    if mode == "pandas":
        import pandas as pd

        df = pd.read_excel(book, sheet_name=sheet_name)
        df = df.dropna(how="all")
        return df.to_dict(orient="records"), list(df.columns)

    ws = book[sheet_name]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    while rows and all(_is_blank(cell) for cell in rows[0]):
        rows.pop(0)
    if not rows:
        return [], []
    header = list(rows[0])
    records: List[Dict[str, Any]] = []
    for row in rows[1:]:
        if all(_is_blank(cell) for cell in row):
            continue
        record: Dict[str, Any] = {}
        for idx, col in enumerate(header):
            if _is_blank(col):
                continue
            record[col] = row[idx] if idx < len(row) else None
        records.append(record)
    return records, header


#=========================================================================
# 1. DATA EXTRACTION & PREPARATION
#=========================================================================

#-------------------------------------------------------------------------
# 1.1. Core data structures
#-------------------------------------------------------------------------

def _select_column(columns: Iterable[Any], candidates: Iterable[str]) -> str | None:
    mapping = {_normalize_key(col): col for col in columns if col is not None}
    for candidate in candidates:
        key = _normalize_key(candidate)
        if key in mapping:
            return mapping[key]
    return None


def _rank_columns(columns: Iterable[Any]) -> List[Any]:
    rank_cols = [col for col in columns if _normalize_key(col).startswith("rank")]

    def sort_key(col: Any) -> Tuple[int, str]:
        text = str(col)
        digits = "".join(ch for ch in text if ch.isdigit())
        return (int(digits) if digits else 9999, text)

    return sorted(rank_cols, key=sort_key)


def _to_int(value: Any, label: str) -> int:
    if _is_blank(value):
        raise ValueError(f"Missing numeric value for {label}.")
    if isinstance(value, bool):
        raise ValueError(f"Invalid boolean value for {label}.")
    if isinstance(value, str):
        value = value.strip()
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid numeric value for {label}: {value}") from exc
    if math.isnan(number):
        raise ValueError(f"Invalid numeric value for {label}: {value}")
    if abs(number - round(number)) > 1e-6:
        logger.warning("Non-integer capacity for {}: {}. Rounding.", label, value)
    return int(round(number))


#-------------------------------------------------------------------------
# 1.2. Classifications and subsets
#-------------------------------------------------------------------------

def _dedupe_preserve(values: List[str], context: str) -> List[str]:
    seen = set()
    output = []
    for val in values:
        if val in seen:
            logger.warning("Duplicate preference '{}' for {}. Ignoring.", val, context)
            continue
        seen.add(val)
        output.append(val)
    return output


#-------------------------------------------------------------------------
# 1.3. Parameters and derived attributes
#-------------------------------------------------------------------------

def load_input_data(source: Any) -> Dict[str, Any]:
    mode, book = _load_workbook(source)

    sheet_names = book.sheet_names if mode == "pandas" else book.sheetnames
    missionarios_sheet = _find_sheet(sheet_names, MISSIONARIOS_SHEETS)
    valencias_sheet = _find_sheet(sheet_names, VALENCIAS_SHEETS)

    mission_records, mission_columns = _read_sheet_records(
        mode, book, missionarios_sheet
    )
    cap_records, cap_columns = _read_sheet_records(mode, book, valencias_sheet)

    name_col = _select_column(mission_columns, ["Nome"])
    if not name_col:
        raise ValueError("Missing 'Nome' column in Missionarios sheet.")

    fixed_col = _select_column(
        mission_columns,
        ["Valencia Fixa", "Alocacao Fixa", "Alocacoes Fixas", "Fixas"],
    )

    rank_cols = _rank_columns(mission_columns)
    if not rank_cols:
        raise ValueError("No Rank columns found in Missionarios sheet.")

    val_col = _select_column(cap_columns, ["Valencia"])
    cap_col = _select_column(
        cap_columns,
        ["Capacidade", "Nº Missionários", "Nº Missionarios", "Capacity"],
    )
    if not val_col or not cap_col:
        raise ValueError("Missing required columns in Valencias sheet.")

    people: List[str] = []
    preferences: Dict[str, List[str]] = {}
    rankings: Dict[str, Dict[str, int]] = {}
    fixed_assignments: Dict[str, str] = {}
    seen_people = set()

    for record in mission_records:
        raw_name = record.get(name_col)
        name = _normalize_label(raw_name)
        if not name:
            logger.warning("Skipping row with missing Nome.")
            continue
        if name in seen_people:
            raise ValueError(f"Duplicate missionario name: {name}")
        seen_people.add(name)

        if fixed_col:
            raw_fixed = record.get(fixed_col)
            if not _is_blank(raw_fixed):
                fixed_valencia = _normalize_label(raw_fixed)
                if name in fixed_assignments:
                    raise ValueError(f"Duplicate fixed allocation for {name}.")
                fixed_assignments[name] = fixed_valencia

        ranked_vals: List[str] = []
        for col in rank_cols:
            val = record.get(col)
            if _is_blank(val):
                continue
            ranked_vals.append(_normalize_label(val))
        ranked_vals = _dedupe_preserve(ranked_vals, name)

        preferences[name] = ranked_vals
        rankings[name] = {val: idx + 1 for idx, val in enumerate(ranked_vals)}
        people.append(name)

        if not ranked_vals:
            logger.warning("Missionario '{}' has no ranked preferences.", name)

    capacities: Dict[str, int] = {}
    for record in cap_records:
        raw_val = record.get(val_col)
        valencia = _normalize_label(raw_val)
        if not valencia:
            continue
        cap_value = _to_int(record.get(cap_col), f"Capacidade for {valencia}")
        if valencia in capacities:
            raise ValueError(f"Duplicate valencia in Valencias: {valencia}")
        capacities[valencia] = cap_value

    if not capacities:
        raise ValueError("No valencias found in Valencias sheet.")

    unknown_valencias = set()
    for name, ranked_vals in preferences.items():
        for val in ranked_vals:
            if val not in capacities:
                unknown_valencias.add(val)
    if unknown_valencias:
        raise ValueError(
            "Unknown valencias referenced in preferences: "
            + ", ".join(sorted(unknown_valencias))
        )

    ranked_set = {val for ranked_vals in preferences.values() for val in ranked_vals}
    for valencia in capacities:
        if valencia not in ranked_set:
            logger.warning("Valencia '{}' is never ranked by any missionario.", valencia)

    total_capacity = sum(capacities.values())
    if total_capacity != len(people):
        logger.error(
            "Capacity mismatch: total capacity {} vs {} missionarios.",
            total_capacity,
            len(people),
        )
        raise SystemExit(1)

    if fixed_assignments:
        unknown_valencias_fixed = sorted(
            val for val in fixed_assignments.values() if val not in capacities
        )
        if unknown_valencias_fixed:
            raise ValueError(
                "Fixed allocations contain unknown valencias: "
                + ", ".join(unknown_valencias_fixed)
            )

        fixed_counts: Dict[str, int] = {val: 0 for val in capacities}
        for val in fixed_assignments.values():
            fixed_counts[val] += 1
        overflow = [
            val for val, count in fixed_counts.items() if count > capacities[val]
        ]
        if overflow:
            raise ValueError(
                "Fixed allocations exceed capacity for: " + ", ".join(overflow)
            )

        for name, val in fixed_assignments.items():
            if val not in preferences.get(name, []):
                logger.warning(
                    "Fixed allocation for '{}' is not in their ranked preferences.",
                    name,
                )

    return {
        "people": people,
        "valencias": list(capacities.keys()),
        "capacities": capacities,
        "preferences": preferences,
        "rankings": rankings,
        "fixed_assignments": fixed_assignments,
        "columns": {
            "person": name_col,
            "valencia": val_col,
        },
        "sheets": {
            "missionarios": missionarios_sheet,
            "valencias": valencias_sheet,
        },
    }
